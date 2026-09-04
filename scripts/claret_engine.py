#!/usr/bin/env python3
"""
claret_engine.py
================
Core shared engine for the claret-version-control-system suite:
- Environment and configuration loader (.env)
- PascalCase file renaming with acronym preservation
- Execution of standalone claret-generator.jar
- Git and GitHub operations (Commit, Push, Tag, Release)
- Release / Tag artifact and source tree downloader
- CLARET DSL parser & locale translator (preserving language tokens)
- Text normalization & CSV diff generator for src/ and output/
"""

import os
import re
import sys
import csv
import json
import shutil
import logging
import zipfile
import tempfile
import unicodedata
import subprocess
from pathlib import Path
from typing import List, Dict, Tuple, Optional, Any
from dotenv import load_dotenv

# Load .env configuration
load_dotenv()

# Logging configuration
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
logger = logging.getLogger("claret_engine")

# ------------------------------------------------------------------------------
# Known Technical Acronyms (to preserve in uppercase)
# ------------------------------------------------------------------------------
KNOWN_ACRONYMS = {
    "CRUD", "LLT", "DSL", "API", "XML", "JSON", "CSV", "HTML", "HTTP", "HTTPS",
    "SQL", "DB", "UI", "UX", "ID", "UUID", "SAFF", "PJE", "MBT", "TGF", "ALTS",
    "XLSX", "DOCX", "ODT", "TXT", "PDF", "REST", "SOAP", "URL", "URI", "DTO",
    "DAO", "GT", "GTP", "ART", "TC", "UC", "SPLAB", "UFCG", "CLARET"
}

# ------------------------------------------------------------------------------
# Configuration Helpers
# ------------------------------------------------------------------------------
def get_config(key: str, default: Optional[str] = None) -> Optional[str]:
    """Retrieve environment variable or fallback to default."""
    return os.getenv(key, default)

def get_github_token() -> Optional[str]:
    """Retrieve GitHub token from environment."""
    return os.getenv("GITHUB_TOKEN")

def get_default_repo() -> Optional[str]:
    """Retrieve default GitHub repository (owner/repo)."""
    return os.getenv("GITHUB_REPO", "diegoquirino/openscience")

def get_default_branch() -> str:
    """Retrieve default working branch."""
    return os.getenv("GITHUB_BRANCH", "claret-version-control-system")

def get_java_cmd() -> str:
    """Resolve Java 26+ runtime command, prioritizing explicitly configured modern JDKs."""
    java_env = os.getenv("JAVA_CMD")
    if java_env and java_env.strip() != "java" and Path(java_env).exists():
        return java_env

    # Search for modern JDKs (Java 26 / 21) before falling back to system PATH
    common_jdk_paths = [
        Path(r"C:\Program Files\Java\jdk-26.0.1\bin\java.exe"),
        Path(r"C:\Program Files\Java\latest\bin\java.exe"),
        Path(r"C:\Program Files\Java\jdk-26\bin\java.exe"),
        Path(r"C:\Program Files\Eclipse Adoptium\jdk-21.0.12.101-hotspot\bin\java.exe"),
        Path("/usr/lib/jvm/java-26-openjdk/bin/java"),
        Path("/usr/lib/jvm/java-21-openjdk/bin/java"),
    ]
    for p in common_jdk_paths:
        if p.exists() and p.is_file():
            return str(p)

    java_home = os.getenv("JAVA_HOME")
    if java_home:
        candidate = Path(java_home) / "bin" / ("java.exe" if sys.platform == "win32" else "java")
        if candidate.exists():
            return str(candidate)

    return java_env if java_env else "java"

def get_claret_jar_path() -> Path:
    """Resolve the path to claret-generator.jar."""
    jar_env = os.getenv("CLARET_JAR_PATH")
    candidates = []
    if jar_env:
        candidates.append(Path(jar_env))
        candidates.append(Path(__file__).resolve().parent.parent / jar_env)
    
    # Common workspace locations
    base_dir = Path(__file__).resolve().parent.parent
    candidates.extend([
        base_dir / "bin" / "claret-generator.jar",
        base_dir / "lib" / "claret-generator.jar",
        base_dir / "claret-generator.jar",
        base_dir / "target" / "claret-generator.jar",
        base_dir.parent / "claret-generator" / "target" / "claret-generator.jar"
    ])

    for p in candidates:
        if p.exists() and p.is_file():
            return p.resolve()
    
    # Fallback to the first specified path even if not yet compiled
    return candidates[0] if candidates else Path("claret-generator.jar")

# ------------------------------------------------------------------------------
# 1. PascalCase Naming with Acronym Retention
# ------------------------------------------------------------------------------
def to_pascal_case_with_acronyms(name: str) -> str:
    """
    Converts a filename or identifier to PascalCase while preserving acronyms in uppercase.
    Examples:
      - 'CRUD_Cliente.claret' -> 'CRUD_Cliente.claret' / 'CRUDCliente.claret'
      - 'login-minitest-alternative-format-dsl.claret' -> 'LoginMinitestAlternativeFormatDsl.claret'
      - 'extracao_LLT.claret' -> 'ExtracaoLLT.claret'
      - 'all_usecases--GT-.xlsx' -> 'AllUsecases--GT-.xlsx'
    """
    path_obj = Path(name)
    stem = path_obj.stem
    extension = path_obj.suffix

    # Special handling for composite suffixes like .dsl.claret
    if stem.endswith(".dsl"):
        stem = stem[:-4]
        extension = ".dsl" + extension

    # Tokenize by underscores, hyphens, spaces, and camelCase boundaries
    tokens = re.split(r'[-_\s]+', stem)
    processed_tokens = []

    for token in tokens:
        if not token:
            continue
        # Split further if camelCase / acronyms concatenated (e.g. ExtracaoLLT)
        sub_tokens = re.findall(r'[A-Z]+(?=[A-Z][a-z])|[A-Z]?[a-z]+|[A-Z]+|[0-9]+', token)
        if not sub_tokens:
            sub_tokens = [token]

        for st in sub_tokens:
            st_upper = st.upper()
            if st_upper in KNOWN_ACRONYMS:
                processed_tokens.append(st_upper)
            elif st.isupper() and len(st) <= 4:
                # Keep short all-caps as acronyms
                processed_tokens.append(st)
            else:
                # Capitalize first letter, keep rest lowercase
                processed_tokens.append(st.capitalize())

    new_stem = "".join(processed_tokens)
    return new_stem + extension

def format_version(version_raw: Any) -> str:
    """
    Format version number into standard X.Y format.
    E.g. 1 -> '1.0', 2 -> '2.0', '1' -> '1.0', '1.1' -> '1.1', 'v1' -> '1.0'
    """
    v_str = str(version_raw).strip().lstrip("vV")
    if not v_str:
        return "1.0"
    if "." not in v_str and v_str.isdigit():
        return f"{v_str}.0"
    return v_str

def format_branch_title(branch_name: str) -> str:
    """
    Formats a branch name into Title Case while preserving uppercase acronyms.
    E.g. 'saff-study' -> 'Saff Study', 'abc-da-net' -> 'ABC Da Net'
    """
    tokens = re.split(r'[-_\s]+', branch_name)
    formatted = []
    for t in tokens:
        if not t:
            continue
        t_upper = t.upper()
        if t_upper in KNOWN_ACRONYMS:
            formatted.append(t_upper)
        elif t.isupper() and len(t) <= 4:
            formatted.append(t)
        else:
            formatted.append(t.capitalize())
    return " ".join(formatted)

# ------------------------------------------------------------------------------
# 2. Standalone claret-generator.jar Runner
# ------------------------------------------------------------------------------
def run_claret_generator(
    input_dir: Path,
    output_dir: Optional[Path] = None,
    formats: str = "all",
    coverage: str = "gt",
    flat: bool = False,
    java_cmd: Optional[str] = None,
    jar_path: Optional[Path] = None
) -> Tuple[bool, str]:
    """
    Executes claret-generator.jar CLI over a given directory.
    Moves specification files to src/ and generates output/ tree.
    """
    resolved_jar = jar_path or get_claret_jar_path()
    if not resolved_jar.exists():
        msg = f"claret-generator.jar not found at {resolved_jar}. Please compile it with 'mvn package'."
        logger.error(msg)
        return False, msg

    resolved_java = java_cmd or get_java_cmd()
    if output_dir is None:
        output_dir = input_dir / "output"

    cmd = [
        resolved_java,
        "-jar",
        str(resolved_jar),
        "-i",
        str(input_dir),
        "-o",
        str(output_dir),
        "-f",
        formats,
        "-c",
        coverage
    ]
    if flat:
        cmd.append("--flat")

    logger.info(f"Executing: {' '.join(cmd)}")
    try:
        res = subprocess.run(cmd, capture_output=True, text=True, check=True)
        logger.info(f"claret-generator output:\n{res.stdout}")
        return True, res.stdout
    except subprocess.CalledProcessError as e:
        logger.error(f"claret-generator failed with exit code {e.returncode}:\n{e.stderr}")
        return False, e.stderr
    except Exception as e:
        logger.error(f"Error running claret-generator: {e}")
        return False, str(e)

# ------------------------------------------------------------------------------
# 3. Git and GitHub Operations
# ------------------------------------------------------------------------------
class GitHubManager:
    """Manages Git operations and GitHub REST API interactions."""

    def __init__(
        self,
        repo: Optional[str] = None,
        token: Optional[str] = None,
        repo_dir: Optional[Path] = None,
        downloads_dir: Optional[Path] = None
    ):
        self.token = token or get_github_token()
        self.repo = repo or get_default_repo()
        self.repo_dir = Path(repo_dir).resolve() if repo_dir else self._discover_local_repo()
        self.downloads_dir = Path(downloads_dir).resolve() if downloads_dir else self._discover_downloads_dir()
        self._cache: Dict[Tuple[str, str], Optional[str]] = {}
        self._tree_cache: Dict[Tuple[str, str], List[str]] = {}
        self._session = None

    def _discover_local_repo(self) -> Optional[Path]:
        branch = os.getenv("GITHUB_BRANCH", "saff-study")
        candidates = [
            Path.cwd() / branch,
            Path.cwd(),
            Path(__file__).resolve().parents[2] / branch,
            Path(__file__).resolve().parents[1] / branch
        ]
        for c in candidates:
            if (c / ".git").is_dir():
                return c
        return None

    def _discover_downloads_dir(self) -> Optional[Path]:
        dl_env = os.getenv("DOWNLOADS_DIR")
        if dl_env and Path(dl_env).is_dir():
            return Path(dl_env).resolve()
        candidates = [
            Path.cwd() / "downloads",
            Path(__file__).resolve().parents[1] / "downloads",
            Path(__file__).resolve().parents[2] / "downloads"
        ]
        for c in candidates:
            if c.is_dir():
                return c.resolve()
        return None

    def _find_downloads_ref_dir(self, ref: str) -> Optional[Path]:
        """Locates downloaded version folder corresponding to a tag or version name."""
        if not self.downloads_dir or not self.downloads_dir.is_dir():
            return None

        # 1. Exact match
        exact = self.downloads_dir / ref
        if exact.is_dir():
            return exact

        ref_clean = ref.strip()
        ref_lower = ref_clean.lower()
        subdirs = [d for d in self.downloads_dir.iterdir() if d.is_dir()]

        # 2. Case-insensitive exact match
        for d in subdirs:
            if d.name.lower() == ref_lower:
                return d

        # 3. Suffix or stripped matching (e.g. saff-study_v1.0 <-> v1.0 <-> 1.0)
        for d in subdirs:
            d_lower = d.name.lower()
            if d_lower.endswith(f"_{ref_lower}") or d_lower.endswith(f"-{ref_lower}"):
                return d
            if not ref_lower.startswith("v") and (d_lower.endswith(f"_v{ref_lower}") or d_lower.endswith(f"-v{ref_lower}")):
                return d
            if "_" in ref_lower:
                suffix = ref_lower.split("_")[-1]
                if d_lower == suffix or d_lower == suffix.lstrip("v"):
                    return d
            if "-" in ref_lower:
                suffix = ref_lower.split("-")[-1]
                if d_lower == suffix or d_lower == suffix.lstrip("v"):
                    return d

        return None

    @staticmethod
    def read_xlsx_as_text(source) -> str:
        """Extracts textual row-by-row representation from an XLSX file path or bytes for diffing."""
        try:
            import io
            import openpyxl
            if isinstance(source, (bytes, bytearray)):
                wb = openpyxl.load_workbook(io.BytesIO(source), data_only=True)
            else:
                wb = openpyxl.load_workbook(source, data_only=True)
            lines = []
            for sheet in wb.worksheets:
                lines.append(f"[Sheet: {sheet.title}]")
                for row in sheet.iter_rows(values_only=True):
                    row_vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
                    if row_vals:
                        lines.append("\t".join(row_vals))
            return "\n".join(lines)
        except Exception as e:
            logger.warning(f"Failed to parse xlsx text: {e}")
            return ""

    @staticmethod
    def read_docx_as_text(source) -> str:
        """Extracts textual representation from a DOCX file path or bytes."""
        try:
            import io
            import zipfile
            zfile = io.BytesIO(source) if isinstance(source, (bytes, bytearray)) else source
            if zipfile.is_zipfile(zfile):
                with zipfile.ZipFile(zfile) as z:
                    if "word/document.xml" in z.namelist():
                        xml_data = z.read("word/document.xml").decode("utf-8", errors="ignore")
                        cleaned = re.sub(r'<[^>]+>', ' ', xml_data)
                        return re.sub(r'[ \t]+', ' ', cleaned).strip()
            return ""
        except Exception as e:
            logger.warning(f"Failed to parse docx text: {e}")
            return ""

    def _local_repo_has_ref(self, ref: str) -> bool:
        if not self.repo_dir or not (self.repo_dir / ".git").is_dir():
            return False
        code, out, _ = self.run_git(["tag", "-l", ref], cwd=self.repo_dir)
        if code == 0 and out.strip() == ref:
            return True
        code, out, _ = self.run_git(["rev-parse", "--verify", f"refs/tags/{ref}"], cwd=self.repo_dir)
        if code == 0:
            return True
        code, out, _ = self.run_git(["rev-parse", "--verify", ref], cwd=self.repo_dir)
        return code == 0

    def _get_session(self):
        if self._session is None:
            import requests
            from requests.adapters import HTTPAdapter
            from urllib3.util import Retry
            s = requests.Session()
            retries = Retry(total=5, backoff_factor=1, status_forcelist=[429, 500, 502, 503, 504], raise_on_status=False)
            adapter = HTTPAdapter(max_retries=retries, pool_connections=10, pool_maxsize=20)
            s.mount("https://", adapter)
            s.mount("http://", adapter)
            self._session = s
        return self._session

    def run_git(self, args: List[str], cwd: Optional[Path] = None) -> Tuple[int, str, str]:
        """Run a git command in the specified working directory."""
        cmd = ["git"] + args
        logger.debug(f"Git command: {' '.join(cmd)} (cwd={cwd})")
        res = subprocess.run(cmd, capture_output=True, text=True, encoding="utf-8", errors="replace", cwd=cwd)
        stdout = res.stdout.strip() if res.stdout else ""
        stderr = res.stderr.strip() if res.stderr else ""
        return res.returncode, stdout, stderr

    def add_commit_push(self, repo_dir: Path, branch: str, commit_message: str) -> bool:
        """Stage all changes, commit, and push to the specified branch."""
        # Ensure directory is a valid git repository
        rc, _, _ = self.run_git(["rev-parse", "--is-inside-work-tree"], cwd=repo_dir)
        if rc != 0:
            self.run_git(["init"], cwd=repo_dir)
            if self.token:
                remote_url = f"https://x-access-token:{self.token}@github.com/{self.repo}.git"
            else:
                remote_url = f"https://github.com/{self.repo}.git"
            self.run_git(["remote", "remove", "origin"], cwd=repo_dir)
            self.run_git(["remote", "add", "origin", remote_url], cwd=repo_dir)

        # Ensure branch exists and is checked out
        self.run_git(["checkout", "-B", branch], cwd=repo_dir)
        self.run_git(["add", "-A"], cwd=repo_dir)

        # Check if there are changes to commit
        code, out, _ = self.run_git(["status", "--porcelain"], cwd=repo_dir)
        if not out:
            logger.info("No changes to commit.")
        else:
            code, out, err = self.run_git(["commit", "-m", commit_message], cwd=repo_dir)
            if code != 0:
                logger.error(f"Git commit failed: {err}")
                return False

        # Push to remote
        code, out, err = self.run_git(["push", "-u", "origin", branch], cwd=repo_dir)
        if code != 0:
            logger.warning(f"Git push standard attempt: {err}. Attempting with upstream.")
            code, out, err = self.run_git(["push", "--set-upstream", "origin", branch], cwd=repo_dir)
            if code != 0:
                logger.error(f"Git push failed: {err}")
                return False
        return True

    def create_tag(self, repo_dir: Path, tag_name: str, message: str) -> bool:
        """Create and push a git tag."""
        code, _, err = self.run_git(["tag", "-a", tag_name, "-m", message], cwd=repo_dir)
        if code != 0 and "already exists" not in err:
            logger.error(f"Failed to create tag {tag_name}: {err}")
            return False
        
        code, _, err = self.run_git(["push", "origin", tag_name], cwd=repo_dir)
        if code != 0:
            logger.warning(f"Failed to push tag {tag_name}: {err}")
        return True

    def create_release(self, tag_name: str, release_title: str, body: str = "") -> bool:
        """Create a GitHub Release via GitHub REST API."""
        if not self.token:
            logger.warning("No GITHUB_TOKEN provided. Skipping GitHub Release creation via API.")
            return False
        
        import requests
        url = f"https://api.github.com/repos/{self.repo}/releases"
        headers = {
            "Authorization": f"token {self.token}",
            "Accept": "application/vnd.github.v3+json"
        }
        payload = {
            "tag_name": tag_name,
            "name": release_title,
            "body": body or f"Automated release for {release_title}",
            "draft": False,
            "prerelease": False
        }
        resp = requests.post(url, headers=headers, json=payload)
        if resp.status_code in [200, 201]:
            logger.info(f"Created GitHub Release: {release_title} ({tag_name})")
            return True
        elif resp.status_code == 422 and "already_exists" in resp.text:
            logger.info(f"Release {release_title} already exists.")
            return True
        else:
            logger.error(f"Failed to create GitHub release: {resp.status_code} - {resp.text}")
            return False

    def fetch_file_content_at_ref(self, ref: str, file_path: str) -> Optional[str]:
        """Fetch file content from local downloads directory, local git tree, or GitHub repository at a specific tag or ref."""
        cache_key = (ref, file_path)
        if cache_key in self._cache:
            return self._cache[cache_key]

        # 1. Try local downloads directory first (if pre-downloaded or generated)
        ref_dir = self._find_downloads_ref_dir(ref)
        if ref_dir:
            local_file = ref_dir / file_path
            if local_file.is_file():
                ext = local_file.suffix.lower()
                if ext == ".xlsx":
                    content = self.read_xlsx_as_text(local_file)
                elif ext == ".docx":
                    content = self.read_docx_as_text(local_file)
                else:
                    try:
                        content = local_file.read_text(encoding="utf-8", errors="replace")
                    except Exception as e:
                        logger.warning(f"Error reading local file {local_file}: {e}")
                        content = None
                if content is not None:
                    self._cache[cache_key] = content
                    return content

        # 2. Try local git repository if available
        if self._local_repo_has_ref(ref):
            if file_path.lower().endswith(".xlsx") or file_path.lower().endswith(".docx"):
                cmd = ["git", "show", f"{ref}:{file_path}"]
                res = subprocess.run(cmd, capture_output=True, cwd=self.repo_dir)
                if res.returncode == 0:
                    if file_path.lower().endswith(".xlsx"):
                        content = self.read_xlsx_as_text(res.stdout)
                    else:
                        content = self.read_docx_as_text(res.stdout)
                    self._cache[cache_key] = content
                    return content
            else:
                code, out, _ = self.run_git(["show", f"{ref}:{file_path}"], cwd=self.repo_dir)
                if code == 0:
                    self._cache[cache_key] = out
                    return out

        # 3. Remote GitHub fetch with session & retries
        session = self._get_session()
        content = None
        is_binary = file_path.lower().endswith(".xlsx") or file_path.lower().endswith(".docx")

        if self.token:
            url = f"https://api.github.com/repos/{self.repo}/contents/{file_path}?ref={ref}"
            headers = {
                "Authorization": f"token {self.token}",
                "Accept": "application/vnd.github.v3.raw"
            }
            try:
                resp = session.get(url, headers=headers, timeout=30)
                if resp.status_code == 200:
                    if is_binary:
                        content = self.read_xlsx_as_text(resp.content) if file_path.lower().endswith(".xlsx") else self.read_docx_as_text(resp.content)
                    else:
                        content = resp.text
            except Exception as e:
                logger.warning(f"Error fetching {file_path} at {ref} via API: {e}")

        if content is None:
            # Fallback to public raw content
            url = f"https://raw.githubusercontent.com/{self.repo}/{ref}/{file_path}"
            try:
                resp = session.get(url, timeout=30)
                if resp.status_code == 200:
                    if is_binary:
                        content = self.read_xlsx_as_text(resp.content) if file_path.lower().endswith(".xlsx") else self.read_docx_as_text(resp.content)
                    else:
                        content = resp.text
            except Exception as e:
                logger.warning(f"Error fetching {file_path} at {ref} via raw URL: {e}")

        self._cache[cache_key] = content
        return content

    def list_tree_at_ref(self, ref: str, path_prefix: str = "src") -> List[str]:
        """List all file paths under path_prefix at a given ref/tag."""
        cache_key = (ref, path_prefix)
        if cache_key in self._tree_cache:
            return self._tree_cache[cache_key]

        # 1. Try local downloads directory first
        ref_dir = self._find_downloads_ref_dir(ref)
        if ref_dir:
            target_prefix_dir = ref_dir / path_prefix
            if target_prefix_dir.is_dir():
                files = [
                    p.relative_to(ref_dir).as_posix()
                    for p in target_prefix_dir.rglob("*")
                    if p.is_file()
                ]
                if files:
                    logger.info(f"Loaded {len(files)} files for ref '{ref}' from local downloads ({ref_dir.name}/{path_prefix})")
                    self._tree_cache[cache_key] = sorted(files)
                    return sorted(files)

        # 2. Try local git repository if available
        if self._local_repo_has_ref(ref):
            code, out, _ = self.run_git(["ls-tree", "-r", "--name-only", ref, path_prefix], cwd=self.repo_dir)
            if code == 0 and out:
                files = [line.strip().replace("\\", "/") for line in out.splitlines() if line.strip()]
                self._tree_cache[cache_key] = files
                return files

        # 3. Remote GitHub REST API with session & retries
        if not self.token:
            logger.warning("GITHUB_TOKEN missing and ref not in local git/downloads; cannot inspect remote git tree via API.")
            return []

        session = self._get_session()
        url = f"https://api.github.com/repos/{self.repo}/git/trees/{ref}?recursive=1"
        headers = {
            "Authorization": f"token {self.token}",
            "Accept": "application/vnd.github.v3+json"
        }
        try:
            resp = session.get(url, headers=headers, timeout=30)
            if resp.status_code != 200:
                logger.error(f"Error fetching tree for ref {ref}: {resp.status_code}")
                return []

            data = resp.json()
            tree = data.get("tree", [])
            files = []
            for item in tree:
                if item.get("type") == "blob":
                    p = item.get("path", "")
                    if p.startswith(path_prefix):
                        files.append(p)
            self._tree_cache[cache_key] = files
            return files
        except Exception as e:
            logger.error(f"Network error fetching tree for ref {ref}: {e}")
            return []

# ------------------------------------------------------------------------------
# 4. CLARET DSL Token-Safe Translator
# ------------------------------------------------------------------------------
CLARET_RESERVED_KEYWORDS = [
    "system", "usecase", "version", "type", "user", "date", "Creation", "Modification",
    "actor", "preCondition", "basicFlow", "step", "af", "ef", "bfs", "alternative",
    "exception", "postCondition"
]

class ClaretTranslator:
    """
    Translates textual requirements inside .claret files while preserving DSL grammar & keywords.
    """

    def __init__(self, target_locale: str = "en-us"):
        self.target_locale = target_locale.lower()
        self.dictionary: Dict[str, str] = {}
        
        # Load external translation mapping if available
        dict_candidates = [
            Path(__file__).resolve().parent / "full_saff_translations.json",
            Path(__file__).resolve().parent.parent / "full_saff_translations.json",
            Path("full_saff_translations.json")
        ]
        for candidate in dict_candidates:
            if candidate.exists():
                try:
                    self.dictionary = json.loads(candidate.read_text(encoding="utf-8"))
                    logger.debug(f"Loaded {len(self.dictionary)} translations from {candidate}")
                    break
                except Exception as e:
                    logger.warning(f"Failed to load translation dictionary from {candidate}: {e}")

    def clean_garbled(self, text: str) -> str:
        """Clean garbled fragments caused by previous partial replacements."""
        t = text
        subs = [
            ("notme", "nome"), ("notvo", "novo"), ("notva", "nova"), ("notvamente", "novamente"),
            ("diagnotstico", "diagnostico"), ("Diagnotstic", "Diagnostic"), ("not ", "no "),
            ("Not ", "No "), ("menots", "menos"), ("Espanotl", "Espanhol"), ("Españotl", "Español"),
            ("Nao ", "Não "), ("nao ", "não "), ("ja ", "já "), ("usuario", "usuário"),
            ("Usuario", "Usuário"), ("extracao", "extração"), ("extracoes", "extrações"),
            ("Extracao", "Extração"), ("Extracoes", "Extrações"), ("secao", "seção"),
            ("secoes", "seções"), ("Secao", "Seção"), ("Secoes", "Seções"),
            ("criacao", "criação"), ("Criacao", "Criação"), ("edicao", "edição"),
            ("Edicao", "Edição"), ("delecao", "deleção"), ("Delecao", "Deleção")
        ]
        for pt, en in subs:
            t = t.replace(pt, en)
        return t

    def translate_text(self, text: str) -> str:
        """Translate a single natural language string chunk."""
        cleaned = text.strip()
        if not cleaned:
            return text

        # 1. Exact match in dictionary
        if cleaned in self.dictionary:
            return self.dictionary[cleaned]

        cleaned_fixed = self.clean_garbled(cleaned)
        if cleaned_fixed in self.dictionary:
            return self.dictionary[cleaned_fixed]

        # 2. Heuristic word/phrase substitutions
        result = cleaned_fixed
        word_subs = [
            ("exibe ", "displays "), ("Exibe ", "Displays "), ("apresenta ", "displays "),
            ("Apresenta ", "Displays "), ("solicitando confirmação", "requesting confirmation"),
            ("retorna à ", "returns to "), ("retorna a ", "returns to "), ("preenche ", "fills "),
            ("Preenche ", "Fills "), ("clica ", "clicks "), ("Clica ", "Clicks "),
            ("seleciona ", "selects "), ("Seleciona ", "Selects "), ("cancela ", "cancels "),
            ("Cancela ", "Cancels "), ("edita ", "edits "), ("Edita ", "Edits "),
            ("deleta ", "deletes "), ("Deleta ", "Deletes "), ("cria ", "creates "),
            ("Cria ", "Creates "), ("salva ", "saves "), ("Salva ", "Saves "),
            ("mensagem ", "message "), ("Mensagem ", "Message "), ("botão ", "button "),
            ("botões ", "buttons "), ("usuário", "user"), ("usuários", "users"),
            ("Usuário", "User"), ("Usuários", "Users"), ("cliente", "customer"),
            ("clientes", "customers"), ("Cliente", "Customer"), ("Clientes", "Customers"),
            ("documento", "document"), ("documentos", "documents"), ("Documento", "Document"),
            ("Documentos", "Documents"), ("laboratório", "laboratory"), ("laboratórios", "laboratories"),
            ("Laboratório", "Laboratory"), ("Laboratórios", "Laboratories"), ("diagnóstico", "diagnostic"),
            ("diagnósticos", "diagnostics"), ("Diagnóstico", "Diagnostic"), ("Diagnósticos", "Diagnostics"),
            ("seção", "section"), ("seções", "sections"), ("Seção", "Section"),
            ("Seções", "Sections"), ("extração", "extraction"), ("extrações", "extractions"),
            ("Extração", "Extraction"), ("Extrações", "Extractions"), ("servidor", "server"),
            ("Servidor", "Server"), ("extrator", "extractor"), ("Extrator", "Extractor"),
            ("relatório", "report"), ("relatórios", "reports"), ("Relatório", "Report"),
            ("Relatórios", "Reports"), ("gráfico", "graph"), ("gráficos", "graphs"),
            ("Gráfico", "Graph"), ("Gráficos", "Graphs"), ("estatística", "statistic"),
            ("estatísticas", "statistics"), ("Estatística", "Statistic"), ("Estatísticas", "Statistics"),
            ("filtro", "filter"), ("filtros", "filters"), ("Filtro", "Filter"),
            ("Filtros", "Filters"), ("configuração", "setting"), ("configurações", "settings"),
            ("Configuração", "Setting"), ("Configurações", "Settings"), ("não existe ", "there is no "),
            ("Não existe ", "There is no "), ("inválido", "invalid"), ("inválidos", "invalid"),
            ("inválida", "invalid"), ("inválidas", "invalid"), ("válido", "valid"),
            ("válidos", "valid"), ("válida", "valid"), ("válidas", "valid"),
            (" e ", " and "), (" ou ", " or "), (" com ", " with "), (" para ", " for "),
            (" de ", " of "), (" da ", " of the "), (" do ", " of the "), (" das ", " of the "),
            (" dos ", " of the "), (" no ", " in the "), (" na ", " in the "),
            (" nos ", " in the "), (" nas ", " in the "), (" por ", " by "),
            (" pelo ", " by the "), (" pela ", " by the "), (" após ", " after "),
            (" antes ", " before "), (" sem ", " without "), (" todos ", " all "),
            (" todas ", " all "), (" tela ", " page "), (" página ", " page "),
            (" campo ", " field "), (" campos ", " fields "), (" pasta ", " folder "),
            (" pastas ", " folders "), (" arquivo ", " file "), (" arquivos ", " files "),
            (" lista ", " list "), (" dados ", " data "), (" erro ", " error "),
            (" sucesso ", " success "), (" realizado ", " completed "), (" realizada ", " completed "),
            (" realizadas ", " completed "), (" realizados ", " completed ")
        ]
        for pt, en in word_subs:
            result = result.replace(pt, en)

        return result

    def translate_claret_content(self, content: str) -> str:
        """
        Translates .claret content line by line, ensuring reserved keywords and structure are preserved.
        """
        lines = content.splitlines()
        translated_lines = []

        for line in lines:
            # 1. Quoted string translation: "..."
            def replace_quote(match):
                inner_text = match.group(1)
                # Check if it's a metadata keyword that shouldn't be touched
                if inner_text in ["Creation", "Modification"]:
                    return f'"{inner_text}"'
                # If date or version format, leave untouched
                if re.match(r'^\d+(\.\d+)*$', inner_text) or re.match(r'^\d{2}/\d{2}/\d{4}$', inner_text):
                    return f'"{inner_text}"'
                translated = self.translate_text(inner_text)
                return f'"{translated}"'

            # Replace quoted strings
            new_line = re.sub(r'"([^"]*)"', replace_quote, line)
            translated_lines.append(new_line)

        return "\n".join(translated_lines)

# ------------------------------------------------------------------------------
# 5. Diff Normalization and CSV Generation
# ------------------------------------------------------------------------------
def normalize_content(content: Optional[str]) -> str:
    """
    Normalizes content according to the specification:
    - Converted to UTF-8
    - Lowercase
    - Collapses adjacent whitespace characters into a single space
    - Collapses adjacent newlines into a single newline
    """
    if content is None:
        return ""
    # Ensure utf-8 text
    if isinstance(content, bytes):
        text = content.decode("utf-8", errors="replace")
    else:
        text = str(content)

    text = text.lower()
    # Normalize lines: split, collapse internal spaces, remove redundant blank lines
    lines = text.splitlines()
    normalized_lines = []
    for l in lines:
        cleaned_line = re.sub(r'[ \t]+', ' ', l).strip()
        if cleaned_line:
            normalized_lines.append(cleaned_line)

    return "\n".join(normalized_lines)

def extract_system_name(content: str) -> str:
    """Extract system name from .claret file content or generated output test suite artifacts."""
    m = re.search(r'system\s+"([^"]+)"', content, re.IGNORECASE)
    if m:
        return m.group(1)
    m2 = re.search(r'system:\s*([^\t\r\n]+)', content, re.IGNORECASE)
    if m2:
        return m2.group(1).strip()
    return "UnknownSystem"

def get_clause_type(line: str) -> str:
    """Classifies a .claret line by its primary semantic clause keyword."""
    l = line.strip().lower()
    for kw in ["actor ", "precondition ", "postcondition ", "step ", "alternative ", "exception ", "version ", "usecase ", "system ", "basicflow"]:
        if l.startswith(kw):
            return kw.strip()
    return "other"

def split_heterogeneous_chunks(src_lines: List[str], tgt_lines: List[str]) -> List[Tuple[List[str], List[str]]]:
    """
    Partitions a replace hunk containing multiple distinct DSL clause types
    (e.g., actor series followed by preCondition) into separate clause-specific chunks.
    """
    src_types = [get_clause_type(l) for l in src_lines]
    tgt_types = [get_clause_type(l) for l in tgt_lines]

    common_types = set(src_types).intersection(set(tgt_types)) - {"other"}
    if len(common_types) > 1:
        ordered_common = []
        for t in src_types:
            if t in common_types and (not ordered_common or ordered_common[-1] != t):
                ordered_common.append(t)

        tgt_order = []
        for t in tgt_types:
            if t in common_types and (not tgt_order or tgt_order[-1] != t):
                tgt_order.append(t)

        if ordered_common == tgt_order and len(ordered_common) > 1:
            sub_chunks = []
            s_idx = 0
            t_idx = 0
            for ct in ordered_common:
                s_part = []
                while s_idx < len(src_lines) and src_types[s_idx] == ct:
                    s_part.append(src_lines[s_idx])
                    s_idx += 1
                t_part = []
                while t_idx < len(tgt_lines) and tgt_types[t_idx] == ct:
                    t_part.append(tgt_lines[t_idx])
                    t_idx += 1
                if s_part or t_part:
                    sub_chunks.append((s_part, t_part))
            if s_idx < len(src_lines) or t_idx < len(tgt_lines):
                sub_chunks.append((src_lines[s_idx:], tgt_lines[t_idx:]))
            return sub_chunks

    return [(src_lines, tgt_lines)]

def extract_granular_diffs(
    raw_src: Optional[str],
    raw_tgt: Optional[str],
    origin_version: str,
    target_version: str,
    file_name: str,
    system_name: str,
    is_dsl: bool = True
) -> List[Dict[str, Any]]:
    """
    Extracts fine-grained diff chunks between two versions of a specification or test case file.
    Instead of serializing the entire file, each added, deleted, or modified block forms a record.
    """
    import difflib
    norm_src = normalize_content(raw_src)
    norm_tgt = normalize_content(raw_tgt)

    if norm_src == norm_tgt:
        return []

    records = []

    # Case 1: Added file
    if not norm_src and norm_tgt:
        records.append({
            "file": file_name,
            "system": system_name,
            "origin_version": origin_version,
            "origin_content": "",
            "target_version": target_version,
            "target_content": norm_tgt
        })
        return records

    # Case 2: Deleted file
    if norm_src and not norm_tgt:
        records.append({
            "file": file_name,
            "system": system_name,
            "origin_version": origin_version,
            "origin_content": norm_src,
            "target_version": target_version,
            "target_content": ""
        })
        return records

    src_lines = norm_src.splitlines()
    tgt_lines = norm_tgt.splitlines()

    sm = difflib.SequenceMatcher(None, src_lines, tgt_lines)

    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            continue

        s_chunk = src_lines[i1:i2]
        t_chunk = tgt_lines[j1:j2]

        if tag == "replace" and is_dsl:
            sub_chunks = split_heterogeneous_chunks(s_chunk, t_chunk)
            for s_sub, t_sub in sub_chunks:
                records.append({
                    "file": file_name,
                    "system": system_name,
                    "origin_version": origin_version,
                    "origin_content": "\n".join(s_sub) if s_sub else "",
                    "target_version": target_version,
                    "target_content": "\n".join(t_sub) if t_sub else ""
                })
        else:
            records.append({
                "file": file_name,
                "system": system_name,
                "origin_version": origin_version,
                "origin_content": "\n".join(s_chunk) if s_chunk else "",
                "target_version": target_version,
                "target_content": "\n".join(t_chunk) if t_chunk else ""
            })

    return records

def generate_diff_csv(
    diff_records: List[Dict[str, Any]],
    output_csv_path: Path
) -> Path:
    """
    Writes diff records to CSV matching the format:
    | # | file | system | origin_version | origin_content | target_version | target_content |
    """
    output_csv_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["#", "file", "system", "origin_version", "origin_content", "target_version", "target_content"]

    with open(output_csv_path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for idx, rec in enumerate(diff_records, start=1):
            writer.writerow({
                "#": idx,
                "file": rec.get("file", ""),
                "system": rec.get("system", ""),
                "origin_version": rec.get("origin_version") or rec.get("source_version", ""),
                "origin_content": rec.get("origin_content") or rec.get("source_content", ""),
                "target_version": rec.get("target_version", ""),
                "target_content": rec.get("target_content", "")
            })
    logger.info(f"Diff CSV generated at: {output_csv_path} with {len(diff_records)} records.")
    return output_csv_path

# ------------------------------------------------------------------------------
# 6. Change Impact Analysis (CIA) Engine: src/ -> output/ (.txt)
# ------------------------------------------------------------------------------
def parse_txt_test_suite(source: Any) -> Dict[str, Any]:
    """
    Parses a generated MBT test suite in .txt format (file Path or raw string) into:
    - system
    - usecase
    - testcases: list of dicts with tc_id, description, flow_type, flow_number, steps, raw_block
    """
    if isinstance(source, Path):
        if not source.exists() or not source.is_file():
            return {"system": "", "usecase": "", "testcases": []}
        content = source.read_text(encoding="utf-8", errors="replace")
    elif isinstance(source, str):
        content = source
    else:
        return {"system": "", "usecase": "", "testcases": []}

    system = ""
    usecase = ""
    m_sys = re.search(r'System:\s*([^\t\r\n]+)', content)
    if m_sys:
        system = m_sys.group(1).strip()
    m_uc = re.search(r'Use Case:\s*([^\t\r\n]+)', content)
    if m_uc:
        usecase = m_uc.group(1).strip()

    tc_blocks = re.split(r'(?=Test Case ID:\s*TC\d+)', content)
    testcases = []

    for block in tc_blocks:
        m_id = re.search(r'Test Case ID:\s*(TC\d+)', block)
        if not m_id:
            continue
        tc_id = m_id.group(1)

        desc = ""
        m_desc = re.search(r'Description:\s*([^\t\r\n]+)', block)
        if m_desc:
            desc = m_desc.group(1).strip()

        flow_type = "basic"
        flow_number = 0
        desc_lower = desc.lower()
        if "alternative flow" in desc_lower:
            flow_type = "alternative"
            m_num = re.search(r'alternative\s+flow\s+(\d+)', desc_lower)
            if m_num:
                flow_number = int(m_num.group(1))
        elif "exception flow" in desc_lower:
            flow_type = "exception"
            m_num = re.search(r'exception\s+flow\s+(\d+)', desc_lower)
            if m_num:
                flow_number = int(m_num.group(1))

        steps = []
        for line in block.splitlines():
            parts = line.split("\t")
            if len(parts) >= 4 and parts[0].strip().isdigit():
                step_num = int(parts[0].strip())
                action = parts[1].strip()
                expected = parts[3].strip() if len(parts) > 3 else ""
                steps.append({
                    "step_num": step_num,
                    "action": action,
                    "expected": expected,
                    "raw": line.strip()
                })

        testcases.append({
            "tc_id": tc_id,
            "description": desc,
            "flow_type": flow_type,
            "flow_number": flow_number,
            "steps": steps,
            "raw_block": block
        })

    return {
        "system": system,
        "usecase": usecase,
        "testcases": testcases
    }

def strip_accents(s: str) -> str:
    """Removes diacritics / accents from text for canonical semantic comparison."""
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def normalize_text_for_search(text: str) -> str:
    """Normalizes text by removing accents, special characters, and collapsing whitespace."""
    s = strip_accents(text.lower())
    s = re.sub(r'[^\w\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def find_enclosing_flow_in_claret(claret_content: str, search_snippet: str) -> Tuple[str, Optional[int], Optional[str]]:
    """
    Finds whether a snippet belongs to basicFlow, alternative <N>, or exception <N>.
    Returns (flow_type, flow_number, flow_name).
    """
    if not claret_content or not search_snippet:
        return ("unknown", None, None)

    norm_snippet = normalize_text_for_search(search_snippet)
    if not norm_snippet:
        return ("unknown", None, None)

    lines = claret_content.splitlines()
    current_flow = "unknown"
    current_num = None
    current_name = None

    for line in lines:
        l_strip = line.strip()
        m_alt = re.match(r'alternative\s+(\d+)\s*,\s*\"([^\"]*)\"', l_strip, re.I)
        if m_alt:
            current_flow = "alternative"
            current_num = int(m_alt.group(1))
            current_name = m_alt.group(2)
        elif re.match(r'exception\s+(\d+)\s*,\s*\"([^\"]*)\"', l_strip, re.I):
            m_exc = re.match(r'exception\s+(\d+)\s*,\s*\"([^\"]*)\"', l_strip, re.I)
            current_flow = "exception"
            current_num = int(m_exc.group(1))
            current_name = m_exc.group(2)
        elif re.match(r'basicFlow\s*\{?', l_strip, re.I):
            current_flow = "basic"
            current_num = 0
            current_name = "Basic Flow"

        norm_line = normalize_text_for_search(line)
        if norm_line and (norm_snippet in norm_line or (len(norm_line) >= 10 and norm_line in norm_snippet)):
            return (current_flow, current_num, current_name)

    return ("unknown", None, None)

def analyze_diff_impact(
    diff_row: Dict[str, Any],
    gh_manager: GitHubManager
) -> Dict[str, Any]:
    """
    Performs Thesis-Aligned Change Impact Analysis (CIA) based on Chapter 1 of the doctoral thesis:
    - 2x2 Matrix Classification:
        * Low actual / Low predicted   -> True Positive  (retain/minor updates)
        * Low actual / High predicted  -> False Negative (needless deletion of good test)
        * High actual / Low predicted  -> False Positive (major change mistaken as minor -> reused test no longer valid)
        * High actual / High predicted -> True Negative  (Yes, discard / major update)
    - 8-Operation Taxonomy: Keep, Update, Remove, Create, Merge, Split, Reassign, Flag
    - 4-Primitive Taxonomy: Retain, Modify, Create, Discard
    """
    file_name = diff_row.get("file", "")
    stem = Path(file_name).stem
    orig_ver = str(diff_row.get("origin_version", "")).strip()
    tgt_ver = str(diff_row.get("target_version", "")).strip()
    orig_content = str(diff_row.get("origin_content", "") or "").strip()
    tgt_content = str(diff_row.get("target_content", "") or "").strip()

    # Retrieve output test suites (.txt) for origin and target versions
    raw_orig_txt = gh_manager.fetch_file_content_at_ref(orig_ver, f"output/txt/{stem}--GT-.txt")
    raw_tgt_txt = gh_manager.fetch_file_content_at_ref(tgt_ver, f"output/txt/{stem}--GT-.txt")

    orig_suite = parse_txt_test_suite(raw_orig_txt or "")
    tgt_suite = parse_txt_test_suite(raw_tgt_txt or "")

    usecase_name = tgt_suite["usecase"] or orig_suite["usecase"] or stem
    combined_content = (orig_content + "\n" + tgt_content).lower()
    o_norm = normalize_text_for_search(orig_content)
    t_norm = normalize_text_for_search(tgt_content)

    # 1. Identify enclosing flow and model element
    is_metadata = all(
        re.match(r'^(?:version\s+|type\s*:|user\s*:|date\s*:|actor\s+)', l.strip(), re.I)
        for l in (orig_content + "\n" + tgt_content).splitlines() if l.strip()
    )
    is_uc_decl = bool(re.search(r'^\s*usecase\s+"', orig_content, re.M) or re.search(r'^\s*usecase\s+"', tgt_content, re.M))

    m_alt = re.search(r'alternative\s+(\d+)', combined_content)
    m_exc = re.search(r'exception\s+(\d+)', combined_content)
    alt_nums = [int(x) for x in re.findall(r'alternative\s+(\d+)', combined_content)]
    exc_nums = [int(x) for x in re.findall(r'exception\s+(\d+)', combined_content)]
    m_alt_hdr = bool(re.search(r'^\s*alternative\s+\d+', orig_content, re.M) or re.search(r'^\s*alternative\s+\d+', tgt_content, re.M))
    m_exc_hdr = bool(re.search(r'^\s*exception\s+\d+', orig_content, re.M) or re.search(r'^\s*exception\s+\d+', tgt_content, re.M))
    flow_type = None
    flow_num = None

    if alt_nums:
        flow_type = "alternative"
        flow_num = alt_nums[0]
    elif exc_nums:
        flow_type = "exception"
        flow_num = exc_nums[0]

    if not flow_type and ("precondition" not in combined_content and "postcondition" not in combined_content) and not is_uc_decl and not is_metadata:
        raw_orig_claret = gh_manager.fetch_file_content_at_ref(orig_ver, f"src/{file_name}") or ""
        raw_tgt_claret = gh_manager.fetch_file_content_at_ref(tgt_ver, f"src/{file_name}") or ""
        search_sample_orig = orig_content.splitlines()[0] if orig_content else ""
        search_sample_tgt = tgt_content.splitlines()[0] if tgt_content else ""

        enc_flow = "unknown"
        enc_num = None
        if search_sample_orig:
            enc_flow, enc_num, _ = find_enclosing_flow_in_claret(raw_orig_claret, search_sample_orig)
        if enc_flow == "unknown" and search_sample_tgt:
            enc_flow, enc_num, _ = find_enclosing_flow_in_claret(raw_tgt_claret, search_sample_tgt)

        if enc_flow in ["alternative", "exception"]:
            flow_type = enc_flow
            flow_num = enc_num
            if enc_flow == "alternative" and enc_num is not None and enc_num not in alt_nums:
                alt_nums.append(enc_num)
            elif enc_flow == "exception" and enc_num is not None and enc_num not in exc_nums:
                exc_nums.append(enc_num)
        elif enc_flow == "basic":
            flow_type = "basic"
            flow_num = 0

    # Branching changes detection (af, ef, bfs)
    af_o = re.findall(r'af:\s*\[([^\]]*)\]', orig_content)
    af_t = re.findall(r'af:\s*\[([^\]]*)\]', tgt_content)
    ef_o = re.findall(r'ef:\s*\[([^\]]*)\]', orig_content)
    ef_t = re.findall(r'ef:\s*\[([^\]]*)\]', tgt_content)
    bfs_o = re.findall(r'bfs:\s*(\d+)', orig_content)
    bfs_t = re.findall(r'bfs:\s*(\d+)', tgt_content)
    branch_changed = bool((af_o != af_t) or (ef_o != ef_t) or (bfs_o != bfs_t))

    # Step actor change detection
    m_act_o = re.search(r'step\s+\d+\s*,\s*([^,]+)\s*,', orig_content)
    m_act_t = re.search(r'step\s+\d+\s*,\s*([^,]+)\s*,', tgt_content)
    actor_changed = bool(m_act_o and m_act_t and m_act_o.group(1).strip().lower() != m_act_t.group(1).strip().lower())

    # 2. Determine Model Element Category
    if is_metadata and combined_content:
        model_element = "MODEL_METADATA"
    elif not orig_content or not tgt_content:
        if is_uc_decl:
            model_element = "USECASE_LIFECYCLE"
        elif "precondition" in combined_content or "postcondition" in combined_content:
            model_element = "PRE_POST_CONDITION"
        elif flow_type == "alternative":
            model_element = "ALTERNATIVE_FLOW"
        elif flow_type == "exception":
            model_element = "EXCEPTION_FLOW"
        else:
            model_element = "USECASE_LIFECYCLE"
    elif is_uc_decl:
        model_element = "USECASE_DECLARATION"
    elif "precondition" in combined_content or "postcondition" in combined_content:
        model_element = "PRE_POST_CONDITION"
    elif branch_changed:
        model_element = "BRANCHING_CONDITION"
    elif actor_changed:
        model_element = "STEP_ACTOR"
    elif flow_type == "basic":
        model_element = "BASIC_FLOW_STEP"
    elif flow_type == "alternative":
        model_element = "ALTERNATIVE_FLOW" if m_alt_hdr else "ALTERNATIVE_FLOW_STEP"
    elif flow_type == "exception":
        model_element = "EXCEPTION_FLOW" if m_exc_hdr else "EXCEPTION_FLOW_STEP"
    else:
        model_element = "OTHER"

    # 3. Determine affected test cases
    affected_orig_cts = []
    affected_tgt_cts = []
    affected_flows_set = set()

    if model_element == "MODEL_METADATA":
        pass
    elif model_element in ["PRE_POST_CONDITION", "USECASE_DECLARATION", "USECASE_LIFECYCLE"]:
        affected_orig_cts = [tc["tc_id"] for tc in orig_suite["testcases"]]
        affected_tgt_cts = [tc["tc_id"] for tc in tgt_suite["testcases"]]
        affected_flows_set.add("All Flows")
    elif flow_type == "alternative":
        nums_to_check = alt_nums if alt_nums else ([flow_num] if flow_num is not None else [])
        for num in nums_to_check:
            affected_flows_set.add(f"Alternative Flow {num}")
        affected_orig_cts = [tc["tc_id"] for tc in orig_suite["testcases"] if tc["flow_type"] == "alternative" and tc["flow_number"] in nums_to_check]
        affected_tgt_cts = [tc["tc_id"] for tc in tgt_suite["testcases"] if tc["flow_type"] == "alternative" and tc["flow_number"] in nums_to_check]
    elif flow_type == "exception":
        nums_to_check = exc_nums if exc_nums else ([flow_num] if flow_num is not None else [])
        for num in nums_to_check:
            affected_flows_set.add(f"Exception Flow {num}")
        affected_orig_cts = [tc["tc_id"] for tc in orig_suite["testcases"] if tc["flow_type"] == "exception" and tc["flow_number"] in nums_to_check]
        affected_tgt_cts = [tc["tc_id"] for tc in tgt_suite["testcases"] if tc["flow_type"] == "exception" and tc["flow_number"] in nums_to_check]
    elif flow_type == "basic":
        actions_to_search = []
        for l in (orig_content + "\n" + tgt_content).splitlines():
            m_act = re.search(r'\"([^\"]+)\"', l)
            if m_act:
                actions_to_search.append(m_act.group(1))

        for tc in orig_suite["testcases"]:
            if tc["flow_type"] == "basic":
                affected_orig_cts.append(tc["tc_id"])
                affected_flows_set.add("Basic Flow")
            else:
                tc_text = tc["raw_block"].lower()
                if any(normalize_text_for_search(act) in normalize_text_for_search(tc_text) for act in actions_to_search):
                    affected_orig_cts.append(tc["tc_id"])
                    affected_flows_set.add(tc["description"].split(":")[0].strip())

        for tc in tgt_suite["testcases"]:
            if tc["flow_type"] == "basic":
                affected_tgt_cts.append(tc["tc_id"])
                affected_flows_set.add("Basic Flow")
            else:
                tc_text = tc["raw_block"].lower()
                if any(normalize_text_for_search(act) in normalize_text_for_search(tc_text) for act in actions_to_search):
                    affected_tgt_cts.append(tc["tc_id"])
                    affected_flows_set.add(tc["description"].split(":")[0].strip())
    else:
        if not orig_content and tgt_content:
            affected_tgt_cts = [tc["tc_id"] for tc in tgt_suite["testcases"]]
            affected_flows_set.add("All Flows")
        elif orig_content and not tgt_content:
            affected_orig_cts = [tc["tc_id"] for tc in orig_suite["testcases"]]
            affected_flows_set.add("All Flows")

    # Total test cases in this usecase across both versions
    all_uc_cts = sorted(list(set([tc["tc_id"] for tc in orig_suite.get("testcases", [])] + [tc["tc_id"] for tc in tgt_suite.get("testcases", [])])), key=lambda x: int(re.sub(r'\D', '', x) or 0))
    total_cts = sorted(list(set(affected_orig_cts + affected_tgt_cts)), key=lambda x: int(re.sub(r'\D', '', x) or 0))
    not_affected_cts_list = [ct for ct in all_uc_cts if ct not in total_cts]
    flows_str = "; ".join(sorted(affected_flows_set)) if affected_flows_set else "N/A"

    # 4. Semantic vs Syntactic Ground Truth Impact Determination
    if is_metadata or (orig_content and tgt_content and o_norm == t_norm):
        actual_impact = "Low"
        tcm_op = "Keep"
        prim_op = "Retain"

        if is_metadata:
            rationale = "Metadados de documentação/versão alterados sem impacto comportamental. O teste deve ser mantido (Keep / Retain), evitando descarte desnecessário."
        else:
            rationale = f"Alteração puramente cosmética/ortográfica que preserva o significado semântico (Exemplo 1 da Tese). O caso de teste deve ser mantido (Keep / Retain): {', '.join(total_cts) if total_cts else 'Nenhum CT afetado'}."

    else:
        actual_impact = "High"

        # Check for specific operations from the 8-operation taxonomy:
        vague_patterns = [r'\bapropriad\w*', r'\bconforme necess\w*', r'\badequad\w*', r'\ba definir\b', r'\bposteriormente\b']
        orig_step_lines = [l for l in orig_content.splitlines() if "step " in l]
        tgt_step_lines = [l for l in tgt_content.splitlines() if "step " in l]

        # 1. Flag (Ambiguity / vague qualifiers introduced - Example 6)
        if any(re.search(pat, tgt_content, re.I) for pat in vague_patterns):
            tcm_op = "Flag"
            prim_op = "Retain"
            rationale = f"Introdução de qualificador vago ou requisito ambíguo (Exemplo 6 da Tese); casos de teste marcados para revisão humana (Flag / Retain pendente): {', '.join(total_cts)}."

        # 2. Reassign (Branching destination bfs changed - Example 5)
        elif "bfs:" in orig_content and "bfs:" in tgt_content and re.search(r'bfs:\s*(\d+)', orig_content) and re.search(r'bfs:\s*(\d+)', tgt_content) and re.search(r'bfs:\s*(\d+)', orig_content).group(1) != re.search(r'bfs:\s*(\d+)', tgt_content).group(1):
            bfs_orig = re.search(r'bfs:\s*(\d+)', orig_content).group(1)
            bfs_tgt = re.search(r'bfs:\s*(\d+)', tgt_content).group(1)
            tcm_op = "Reassign"
            prim_op = "Modify"
            rationale = f"Destino de ramificação alterado de bfs:{bfs_orig} para bfs:{bfs_tgt} (Exemplo 5 da Tese); requer reatribuição de rastreabilidade do fluxo (Reassign / Modify): {', '.join(total_cts)}."

        # 3. Merge (Consolidation of multiple steps/rules into single step - Example 3)
        elif len(orig_step_lines) >= 2 and len(tgt_step_lines) == 1:
            tcm_op = "Merge"
            prim_op = "Modify"
            rationale = f"Múltiplos passos/regras prévios consolidados em um único passo unificado (Exemplo 3 da Tese); casos de teste devem ser unificados (Merge / Modify): {', '.join(total_cts)}."

        # 4. Split (Single step expanded into multiple distinct verifiable steps - Example 4)
        elif len(orig_step_lines) == 1 and len(tgt_step_lines) >= 2:
            tcm_op = "Split"
            prim_op = "Create"
            rationale = f"Passo expandido em múltiplos comportamentos independentemente verificáveis (Exemplo 4 da Tese); requer decomposição em novos testes focados (Split / Create)."

        # 5. Create (Addition of new flow, alternative, or usecase)
        elif not orig_content and tgt_content:
            tcm_op = "Create"
            prim_op = "Create"
            rationale = f"Novo fluxo ou elemento comportamental introduzido na especificação (Exemplo 2 da Tese); requer criação de novos casos de teste (Create / Create): {', '.join(total_cts)}."

        # 6. Remove (Deletion of flow, alternative, or usecase)
        elif orig_content and not tgt_content:
            tcm_op = "Remove"
            prim_op = "Discard"
            rationale = f"Fluxo ou comportamento descontinuado na especificação (Exemplo 2 da Tese); casos de teste obsoletos devem ser descartados (Remove / Discard): {', '.join(total_cts)}."

        # 7. Update (Default behavioral semantic modification)
        else:
            tcm_op = "Update"
            prim_op = "Modify"
            if model_element == "PRE_POST_CONDITION":
                rationale = f"Pré/Pós-condição alterada semanticamente; impacta o estado inicial/final e requer atualização de todos os casos de teste do caso de uso (Update / Modify): {', '.join(total_cts)}."
            elif model_element == "BRANCHING_CONDITION":
                rationale = f"Condição/referência de desvio alterada; modifica a navegação do grafo de fluxos e caminhos de teste associados (Update / Modify): {', '.join(total_cts)}."
            elif model_element == "STEP_ACTOR":
                rationale = f"Ator executor do passo alterado; modifica o contexto de papéis e permissões do teste (Update / Modify): {', '.join(total_cts)}."
            elif model_element == "USECASE_DECLARATION":
                rationale = f"Declaração/nome do caso de uso alterado; impacta a identificação de todos os casos de teste da suíte (Update / Modify): {', '.join(total_cts)}."
            else:
                rationale = f"Modificação semântica comportamental no passo/fluxo (Exemplo 2 da Tese); requer atualização dos casos de teste impactados para refletir o novo comportamento (Update / Modify): {', '.join(total_cts)}."

    return {
        "usecase": usecase_name,
        "actual_change_impact": actual_impact,
        "model_element": model_element,
        "tcm_operation": tcm_op,
        "primitive_operation": prim_op,
        "affected_cts_count": len(total_cts),
        "affected_cts": ", ".join(total_cts),
        "not_affected_cts_count": len(not_affected_cts_list),
        "not_affected_cts": ", ".join(not_affected_cts_list),
        "affected_flows": flows_str,
        "semantic_rationale": rationale
    }

def generate_cia_csv(
    cia_records: List[Dict[str, Any]],
    output_csv_path: Path
) -> Path:
    """
    Writes the Truth Table Change Impact Analysis (CIA) CSV:
    | # | file | system | origin_version | origin_content | target_version | target_content |
      usecase | actual_change_impact | model_element | tcm_operation | primitive_operation |
      affected_cts_count | affected_cts | not_affected_cts_count | not_affected_cts |
      affected_flows | semantic_rationale |
    """
    output_csv_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "#", "file", "system", "origin_version", "origin_content", "target_version", "target_content",
        "usecase", "actual_change_impact", "model_element", "tcm_operation", "primitive_operation",
        "affected_cts_count", "affected_cts", "not_affected_cts_count", "not_affected_cts",
        "affected_flows", "semantic_rationale"
    ]

    with open(output_csv_path, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for idx, rec in enumerate(cia_records, start=1):
            writer.writerow({
                "#": idx,
                "file": rec.get("file", ""),
                "system": rec.get("system", ""),
                "origin_version": rec.get("origin_version") or rec.get("source_version", ""),
                "origin_content": rec.get("origin_content") or rec.get("source_content", ""),
                "target_version": rec.get("target_version", ""),
                "target_content": rec.get("target_content", ""),
                "usecase": rec.get("usecase", ""),
                "actual_change_impact": rec.get("actual_change_impact", ""),
                "model_element": rec.get("model_element", ""),
                "tcm_operation": rec.get("tcm_operation", ""),
                "primitive_operation": rec.get("primitive_operation", ""),
                "affected_cts_count": rec.get("affected_cts_count", 0),
                "affected_cts": rec.get("affected_cts", ""),
                "not_affected_cts_count": rec.get("not_affected_cts_count", 0),
                "not_affected_cts": rec.get("not_affected_cts", ""),
                "affected_flows": rec.get("affected_flows", ""),
                "semantic_rationale": rec.get("semantic_rationale", "")
            })

    logger.info(f"Truth table CIA CSV successfully generated at: {output_csv_path} with {len(cia_records)} records.")
    return output_csv_path


